#! python2
# -*- coding:utf-8 -*-
"""补关键字和微博两列，不断复制上一行
思路1：冒泡法
思路2：先确定行数和列数，再找到不为空的行数，对不为空与不为空之间的空进行处理，相当于列表切片
思路3: 1.循环每一行，找到非空的，并记录非空的行数 2.对于空的同样记录行数 3.用空的行数去减非空的行数，差值小于两个非空之间的差值"""
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl import Workbook

# wb = openpyxl.load_workbook('Weibo-NSX.xlsx')
# sheet = wb.get_sheet_by_name('Worksheet')
# for rowi in range(2, sheet.max_row):
#     for rowj in range(3, sheet.max_row+1):
#         if sheet.cell(row=rowi, column=1).value is None:



# wb = openpyxl.load_workbook('Weibo-NSX.xlsx')
# sheet = wb.get_sheet_by_name('Worksheet')
# # d_value = 0  # 两个有内容单元格之间的差值
# previous = 0  # 前一个有内容的单元格的行数
# after = 0  # 后一个有内容的单元格的行数
# for col_i in sheet.iter_cols(min_row=2, max_col=2, max_row=sheet.max_row-1):
#     for col_j in sheet.iter_cols(min_row=3, max_col=2, max_row=sheet.max_row):
#         if col_i.value:  #  第一个非空
#             previous = col_i.row
#         if col_j.value:  #  第二个非空
#             after = col_j.row
#         d_value = after - previous
#         if col_i.value is None:
#             if col_i.row - previous < d_value:
#                 # col_i.value = copy(previous)


# wb = openpyxl.load_workbook('Weibo-NSX.xlsx')
# sheet = wb.get_sheet_by_name('Worksheet')
# print sheet.max_row
# previous = 0
# after = 0
# for col_i in sheet.iter_cols(min_row=2, max_col=2, max_row=sheet.max_row-1):
#     for col_j in sheet.iter_cols(min_row=3, max_col=2, max_row=sheet.max_row):
#         for cell_i in col_i:
#             for cell_j in col_j:
#                 if cell_i.value:
#                     previous = cell_i.row
#                 if cell_j.value:
#                     after = cell_j.row
#                 d_value = after - previous
#                 if cell_i.value is None:
#                     if cell_i.row - previous < d_value:
#                         cell_i.value = previous.value


# wb = openpyxl.load_workbook('test1.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1')
# previous_row = 0
# previous_column = 0
# after_row = 0
# after_column = 0
# for col_i in sheet.iter_cols(min_row=2, max_col=2, max_row=sheet.max_row-1):
#     for col_j in sheet.iter_cols(min_row=3, max_col=2, max_row=sheet.max_row):
#         for cell_i in col_i:
#             for cell_j in col_j:
#                 if cell_i.value:
#                     previous_row = cell_i.row
#                     previous_column = cell_i.column
#                 if cell_j.value:
#                     after_row = cell_j.row
#                     after_column = cell_j.column
#                 d_value = after_row - previous_row  # 有可能为负数
#                 if d_value > 0:
#                     if cell_i.value is None:
#                         if cell_i.row - previous_row < d_value:
#                             # cell_i.value = sheet.cell(row=previous_row, column=previous_column).value
#                             cell_i.value = sheet['%s%s' % (previous_column, previous_row)].value
# wb.save('test1.xlsx')


# thin_border = Border(left=Side(style='thin'),
#                      right=Side(style='thin'),
#                      top=Side(style='thin'),
#                      bottom=Side(style='thin'))
# wb = openpyxl.load_workbook('test1.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1')
# sheet.cell(row=2, column=2).border = thin_border
# wb.save('test1.xlsx')


# wb = openpyxl.load_workbook('Weibo-NSX.xlsx')
# sheet = wb.get_sheet_by_name('Worksheet')
# # for idx in range(2, sheet.max_row):
# #     sheet['B%s' % idx].alignment = Alignment(wrapText=True)
# for col in sheet.iter_cols(min_row=1, min_col=1, max_col=sheet.max_col, max_row=sheet.max_row):
#     for cell in col:
#         sheet['%s%s' % (cell.column, cell.row)].alignment = Alignment(wrapText=True)
# previous_row = 0
# previous_column = 0
# after_row = 0
# after_column = 0
# for col_i in sheet.iter_cols(min_row=2, max_col=2, max_row=sheet.max_row-1):
#     for col_j in sheet.iter_cols(min_row=3, max_col=2, max_row=sheet.max_row):
#         for cell_i in col_i:
#             for cell_j in col_j:
#                 if cell_i.value:
#                     previous_row = cell_i.row
#                     previous_column = cell_i.column
#                 if cell_j.value:
#                     after_row = cell_j.row
#                     after_column = cell_j.column
#                 d_value = after_row - previous_row  # 有可能为负数
#                 if d_value > 0:
#                     if cell_i.value is None:
#                         if cell_i.row - previous_row < d_value:
#                             cell_i.value = sheet['%s%s' % (previous_column, previous_row)].value
# wb.save('Weibo-NSX.xlsx')


# wb = openpyxl.load_workbook('test.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1')
# # for idx in range(2, sheet.max_row):
# #     sheet['B%s' % idx].alignment = Alignment(wrapText=True)
# for col in sheet.iter_cols(max_col=sheet.max_column, max_row=sheet.max_row):
#     for cell in col:
#         sheet['%s%s' % (cell.column, cell.row)].alignment = Alignment(wrapText=True)
# previous_row = 0
# previous_column = 0
# after_row = 0
# after_column = 0
# for col_i in sheet.iter_cols(min_row=2, max_col=2, max_row=sheet.max_row-1):
#     for col_j in sheet.iter_cols(min_row=3, max_col=2, max_row=sheet.max_row):
#         for cell_i in col_i:
#             for cell_j in col_j:
#                 if cell_i.value:
#                     previous_row = cell_i.row
#                     previous_column = cell_i.column
#                 if cell_j.value:
#                     after_row = cell_j.row
#                     after_column = cell_j.column
#                 d_value = after_row - previous_row  # 有可能为负数
#                 if d_value > 0:
#                     if cell_i.value is None:
#                         if cell_i.row - previous_row < d_value:
#                             cell_i.value = sheet['%s%s' % (previous_column, previous_row)].value
# wb.save('test.xlsx')


wb = openpyxl.load_workbook('Weibo-NSX.xlsx')
sheet = wb.get_sheet_by_name('Worksheet')
for idx in range(2, sheet.max_row):
    sheet['B%s' % idx].alignment = Alignment(shrink_to_fit=True)
# for col in sheet.iter_cols(max_col=sheet.max_column, max_row=sheet.max_row):
#     for cell in col:
#         sheet['%s%s' % (cell.column, cell.row)].alignment = Alignment(wrapText=True, shrink_to_fit=True)
previous_row = 0
previous_column = 0
after_row = 0
after_column = 0
for col_i in sheet.iter_cols(min_row=2, max_col=2, max_row=sheet.max_row-1):
    for col_j in sheet.iter_cols(min_row=3, max_col=2, max_row=sheet.max_row):
        for cell_i in col_i:
            for cell_j in col_j:
                if cell_i.value:
                    previous_row = cell_i.row
                    previous_column = cell_i.column
                if cell_j.value:
                    after_row = cell_j.row
                    after_column = cell_j.column
                d_value = after_row - previous_row  # 有可能为负数
                if d_value > 0:
                    if cell_i.value is None:
                        if cell_i.row - previous_row < d_value:
                            cell_i.value = sheet['%s%s' % (previous_column, previous_row)].value

wb.save('success.xlsx')