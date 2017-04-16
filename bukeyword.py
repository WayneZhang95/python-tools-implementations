#! python2
# -*- coding:utf-8 -*-
"""根据id列去数据库中去找主键,将keyword那一列补到数据库里"""
import openpyxl
from autohome_model import News


wb = openpyxl.load_workbook('news.xlsx')
sheet = wb.get_sheet_by_name('sheet1')
for col in sheet.iter_cols(min_row=2, max_col=1, max_row=sheet.max_row):
    for cell in col:
        if News.select().where(News.id == cell.value).exists():
            excel_keyword = sheet['H%s' % cell.row].value
            u = News.update(keyword=excel_keyword).where(News.id == cell.value)
            u.execute()