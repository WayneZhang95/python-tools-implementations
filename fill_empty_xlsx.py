#! python2
# -*- coding:utf-8 -*-
"""补关键字和微博两列，不断复制上一行
思路1：冒泡法
思路2：先确定行数和列数，再找到不为空的行数，对不为空与不为空之间的空进行处理，相当于列表切片
思路3: 1.循环每一行，找到非空的，并记录非空的行数 2.对于空的同样记录行数 3.用空的行数去减非空的行数，差值小于两个非空之间的差值"""
import openpyxl
from openpyxl.styles import Alignment


wb = openpyxl.load_workbook('Weibo-NSX.xlsx')
sheet = wb.get_sheet_by_name('Worksheet')
for idx in range(2, sheet.max_row):
    sheet['B%s' % idx].alignment = Alignment(shrink_to_fit=True)  # 当xlsx文件格子中字符长度很长时，要用shrink_to_fit属性
previous_row = 0
previous_column = 0
after_row = 0
after_column = 0
for col_i in sheet.iter_cols(min_row=2, max_col=2, max_row=sheet.max_row-1):
    for col_j in sheet.iter_cols(min_row=3, max_col=2, max_row=sheet.max_row):
        for cell_i in col_i:
            for cell_j in col_j:
                if cell_i.value:
                    previous_row = cell_i.row
                    previous_column = cell_i.column
                if cell_j.value:
                    after_row = cell_j.row
                    after_column = cell_j.column
                d_value = after_row - previous_row  # 有可能为负数
                if d_value > 0:
                    if cell_i.value is None:
                        if cell_i.row - previous_row < d_value:
                            cell_i.value = sheet['%s%s' % (previous_column, previous_row)].value

wb.save('success.xlsx')